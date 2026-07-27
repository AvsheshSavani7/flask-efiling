"""
South Africa CompCom MA Weekly Case List → sa_compcom_cases collection
======================================================================
Loads https://www.compcom.co.za/2026-2/, downloads the latest weekly XLSX
(or all weeks in --bootstrap), inserts new cases, matches deals, and
silently marks pending DB cases missing from the current list as
"removed from pending list".

Normal run: if the latest XLSX URL already exists on any DB record
(source_xlsx_url), skip the entire run. Otherwise LLM → regex → USA
check → email; always insert new cases.
--test-email: same as normal, but emails go to TEST_RECIPIENT via
send_direct_email + N8N_WEBHOOK_ONLY_ME (no org routing).
Bootstrap: insert without LLM / regex / USA / email (no URL skip).
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import tempfile
from datetime import datetime, timezone, timedelta
from html import escape as escape_html
from logging.handlers import RotatingFileHandler
from typing import Any, Dict, List, Optional, Set, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from deal_match_llm import fetch_open_deals, llm_match_deal_id
from deal_match_regex import apply_regex_match_subject, regex_match_sa_compcom_deal
from email_subject_builder import build_subject
from llm_verification_service import verify_usa_relation
from log_utils import cleanup_old_logs, refresh_log_file
from mongodb_connection import (
    get_database,
    get_deal_by_id,
    init_mongodb_connection,
    is_connected,
)
from n8n_email_service import post_email_payload, send_direct_email
from scraper_error_utils import collect_error, send_error_summary

load_dotenv(".env")

LIST_PAGE_URL = "https://www.compcom.co.za/2026-2/"
BACKUP_JSON = "sa_compcom_cases_register_backup.json"
ENV_PATH = ".env"
COLLECTION_NAME = "sa_compcom_cases"
REMOVED_STATUS = "removed from pending list"
PENDING_STATUS = "Pending"
TEST_RECIPIENT = "avshesh.savani@teqnodux.com"

PERSISTENT_LOG_DIR = "/var/data/logs"
SCRIPT_NAME = "sa_compcom_cases_register"
IST = timezone(timedelta(hours=5, minutes=30))

LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
LOG_RETENTION_DAYS = int(os.getenv("LOG_RETENTION_DAYS", "30"))
LOG_MAX_BYTES = int(os.getenv("LOG_MAX_BYTES", str(2 * 1024 * 1024)))
LOG_BACKUP_COUNT = int(os.getenv("LOG_BACKUP_COUNT", "3"))

HEADER_ALIASES = {
    "no": "row_no",
    "case number": "case_number",
    "primary acquiring firm": "primary_acquiring_firm",
    "primary target firm": "primary_target_firm",
    "initial date filed": "initial_date_filed",
    "size": "size",
    "due date": "due_date",
    "extension date for decision": "extension_date",
    "phase": "phase",
    "status": "status",
}


def _get_log_file() -> str:
    base = PERSISTENT_LOG_DIR if os.path.isdir("/var/data") else "."
    log_dir = os.path.join(base, SCRIPT_NAME)
    os.makedirs(log_dir, exist_ok=True)
    today = datetime.now(IST).strftime("%Y-%m-%d")
    return os.path.join(log_dir, f"{today}.log")


LOG_FILE = _get_log_file()


class _ISTFormatter(logging.Formatter):
    def converter(self, timestamp):
        return datetime.fromtimestamp(timestamp, tz=IST)

    def formatTime(self, record, datefmt=None):
        ct = self.converter(record.created)
        if datefmt:
            return ct.strftime(datefmt)
        return ct.strftime("%Y-%m-%d %I:%M:%S %p IST")


logger = logging.getLogger(SCRIPT_NAME)
logger.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))

if not logger.handlers:
    formatter = _ISTFormatter(fmt="%(asctime)s | %(levelname)s | %(message)s")
    fh = RotatingFileHandler(
        LOG_FILE,
        maxBytes=LOG_MAX_BYTES,
        backupCount=LOG_BACKUP_COUNT,
        encoding="utf-8",
    )
    fh.setFormatter(formatter)
    logger.addHandler(fh)
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(formatter)
    logger.addHandler(sh)

logger.propagate = False
cleanup_old_logs(os.path.dirname(LOG_FILE), LOG_RETENTION_DAYS)


def utc_now_iso() -> str:
    return (
        datetime.now(timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z")
    )


def get_sa_compcom_cases_collection():
    db = get_database()
    if db is None:
        return None
    return db[COLLECTION_NAME]


def fetch_list_page_html(url: str = LIST_PAGE_URL, headless: bool = True) -> Optional[str]:
    """Load the weekly list page via Playwright; fall back to requests."""
    try:
        logger.info("Fetching CompCom list page via Playwright: %s", url)
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=headless)
            try:
                page = browser.new_page()
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                page.wait_for_timeout(1500)
                html = page.content()
                logger.info("Fetched HTML via Playwright (%s bytes)", len(html))
                return html
            finally:
                browser.close()
    except Exception as e:
        logger.warning("Playwright fetch failed (%s); falling back to requests", e)

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=60)
        resp.raise_for_status()
        logger.info("Fetched HTML via requests (%s bytes)", len(resp.text))
        return resp.text
    except requests.RequestException as e:
        logger.error("Error fetching list page: %s", e)
        return None


def _parse_week_label(label: str) -> Optional[datetime]:
    cleaned = re.sub(r"\s+", " ", (label or "").replace("\xa0", " ")).strip()
    if not cleaned:
        return None
    for fmt in ("%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            continue
    return None


def parse_weekly_xlsx_links(html_content: str, base_url: str = LIST_PAGE_URL) -> List[Dict[str, Any]]:
    """Parse MA Weekly Case List XLSX links from the CompCom year page."""
    soup = BeautifulSoup(html_content, "html.parser")
    links: List[Dict[str, Any]] = []
    seen: Set[str] = set()

    content = soup.select_one(".financity-content-area") or soup
    for a in content.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        if not href.lower().endswith(".xlsx"):
            continue
        if "MA-Weekly-Case-List" not in href and "Weekly-Case-List" not in href:
            # Still accept CompCom upload xlsx under wp-content/uploads
            if "/wp-content/uploads/" not in href:
                continue

        url = urljoin(base_url, href)
        if url in seen:
            continue
        seen.add(url)

        label = a.get_text(" ", strip=True).replace("\xa0", " ").strip()
        week_dt = _parse_week_label(label)
        links.append(
            {
                "url": url,
                "label": label,
                "week_date": week_dt.strftime("%Y-%m-%d") if week_dt else "",
                "week_dt": week_dt,
            }
        )

    links.sort(
        key=lambda item: item["week_dt"] or datetime.min,
        reverse=True,
    )
    logger.info("Parsed %s weekly XLSX links", len(links))
    return links


def download_xlsx(url: str) -> Optional[str]:
    """Download XLSX to a temp file; return path or None."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
    }
    try:
        logger.info("Downloading XLSX: %s", url)
        resp = requests.get(url, headers=headers, timeout=120)
        resp.raise_for_status()
        fd, path = tempfile.mkstemp(prefix="sa_compcom_", suffix=".xlsx")
        os.close(fd)
        with open(path, "wb") as f:
            f.write(resp.content)
        logger.info("Saved XLSX (%s bytes) → %s", len(resp.content), path)
        return path
    except requests.RequestException as e:
        logger.error("Failed to download XLSX %s: %s", url, e)
        return None


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text).strip()
    return text


def _is_header_row(values: List[Any]) -> bool:
    joined = " ".join(_cell_to_str(v).lower() for v in values[:4])
    return "case number" in joined and "acquiring" in joined


def parse_xlsx_rows(
    path: str,
    source_xlsx_url: str,
    list_week_date: str,
) -> List[Dict[str, Any]]:
    """Parse CompCom weekly case list XLSX into row dicts."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_map: Dict[int, str] = {}
    rows: List[Dict[str, Any]] = []

    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if not any(v is not None and str(v).strip() for v in values):
            continue

        if _is_header_row(values) or not header_map:
            if _is_header_row(values):
                header_map = {}
                for idx, cell in enumerate(values):
                    key = _cell_to_str(cell).lower()
                    field = HEADER_ALIASES.get(key)
                    if field:
                        header_map[idx] = field
                continue
            # Title / blank rows before header
            continue

        record: Dict[str, Any] = {
            field: "" for field in HEADER_ALIASES.values()
        }
        for idx, field in header_map.items():
            if idx < len(values):
                record[field] = _cell_to_str(values[idx])

        case_number = record.get("case_number", "").strip()
        if not case_number or case_number.lower() == "case number":
            continue

        record["case_number"] = case_number
        record["source_xlsx_url"] = source_xlsx_url
        record["list_week_date"] = list_week_date
        record["list_page_url"] = LIST_PAGE_URL
        rows.append(record)

    logger.info("Parsed %s case rows from XLSX", len(rows))
    return rows


def case_exists(collection, case_number: str) -> bool:
    try:
        return (
            collection.count_documents({"case_number": case_number}, limit=1) > 0
        )
    except Exception as e:
        logger.exception("Error checking existing case: %s", e)
        return False


def xlsx_url_already_processed(collection, source_xlsx_url: str) -> bool:
    """True if any sa_compcom_cases record was ingested from this XLSX URL."""
    if not source_xlsx_url:
        return False
    try:
        return (
            collection.count_documents(
                {"source_xlsx_url": source_xlsx_url}, limit=1
            )
            > 0
        )
    except Exception as e:
        logger.exception("Error checking source_xlsx_url: %s", e)
        return False


def ensure_indexes(collection) -> None:
    try:
        collection.create_index("case_number", unique=True)
    except Exception as e:
        logger.warning("Could not create unique index on case_number: %s", e)


def match_case_to_deal(
    case_number: str,
    acquiring: str,
    target: str,
    deals: Optional[List[Dict[str, Any]]] = None,
) -> Optional[str]:
    return llm_match_deal_id(
        regulator_name="South Africa Competition Commission",
        case_sections={
            "CASE NUMBER": case_number,
            "PRIMARY ACQUIRING FIRM": acquiring,
            "PRIMARY TARGET FIRM": target,
        },
        source_label="the South Africa CompCom acquiring and target firms",
        deals=deals,
    )


def generate_matched_case_email_html(
    case_info: Dict[str, Any], deal: Dict[str, Any]
) -> str:
    target = deal.get("target") or deal.get("target_name", "N/A")
    acquirer = deal.get("acquirer") or deal.get("acquire_name", "N/A")
    deal_id = str(deal.get("deal_id") or "N/A")
    list_url = case_info.get("list_page_url") or LIST_PAGE_URL

    return f"""<!doctype html>
<html lang="en">
<head><meta charset="utf-8" /><title>South Africa CompCom - New Case</title></head>
<body style="margin:0;padding:0;background:#ffffff;color:#0f172a;font-family:system-ui,-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;">
<div style="max-width:1100px;margin:0 auto;padding:28px 26px 40px 26px;">
<div style="background:#dbeafe;border-radius:6px;padding:16px 22px;margin-bottom:20px;border-left:4px solid #2563eb;">
  <div style="font-size:15px;font-weight:800;color:#1e40af;margin-bottom:6px;">Matched Deal</div>
  <div style="font-size:14px;color:#1e3a8a;">
    <span style="font-weight:700;">Acquirer:</span> {escape_html(str(acquirer))}
    <span style="color:#94a3b8;margin:0 8px;">|</span>
    <span style="font-weight:700;">Target:</span> {escape_html(str(target))}
    <span style="color:#94a3b8;margin:0 8px;">|</span>
    <span style="font-weight:700;">Deal ID:</span> {escape_html(deal_id)}
  </div>
  <div style="margin-top:10px;">
    <a href="{escape_html(list_url)}" target="_blank" style="color:#2563eb;text-decoration:none;font-weight:700;font-size:14px;">View Weekly Case List →</a>
  </div>
</div>
<div style="background:#f3f4f6;border-radius:6px;padding:22px 26px;">
  <div style="font-size:18px;font-weight:800;margin-bottom:16px;">South Africa CompCom - New Case</div>
  <div style="display:grid;grid-template-columns:220px 1fr;row-gap:12px;column-gap:18px;">
    <div style="font-weight:700;">Case Number:</div><div>{escape_html(case_info.get("case_number", "N/A"))}</div>
    <div style="font-weight:700;">Acquiring Firm:</div><div>{escape_html(case_info.get("primary_acquiring_firm", "N/A"))}</div>
    <div style="font-weight:700;">Target Firm:</div><div>{escape_html(case_info.get("primary_target_firm", "N/A"))}</div>
    <div style="font-weight:700;">Initial Date Filed:</div><div>{escape_html(case_info.get("initial_date_filed", "N/A"))}</div>
    <div style="font-weight:700;">Size:</div><div>{escape_html(case_info.get("size", "N/A"))}</div>
    <div style="font-weight:700;">Due Date:</div><div>{escape_html(case_info.get("due_date", "N/A"))}</div>
    <div style="font-weight:700;">Extension Date:</div><div>{escape_html(case_info.get("extension_date", "N/A") or "—")}</div>
    <div style="font-weight:700;">Phase:</div><div>{escape_html(case_info.get("phase", "N/A"))}</div>
    <div style="font-weight:700;">Status:</div><div>{escape_html(case_info.get("status", "N/A"))}</div>
    <div style="font-weight:700;">List Week:</div><div>{escape_html(case_info.get("list_week_date", "N/A"))}</div>
  </div>
</div>
</div>
</body>
</html>"""


def generate_usa_related_email_html(case_info: Dict[str, Any]) -> str:
    list_url = case_info.get("list_page_url") or LIST_PAGE_URL
    return f"""<!doctype html>
<html lang="en">
<head><meta charset="utf-8" /><title>USA-Related South Africa CompCom Case</title></head>
<body style="margin:0;padding:0;background:#ffffff;color:#0f172a;font-family:system-ui,-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;">
<div style="max-width:1100px;margin:0 auto;padding:28px 26px 40px 26px;">
<div style="background:#dbeafe;border-radius:6px;padding:16px 22px;margin-bottom:20px;border-left:4px solid #3b82f6;">
  <div style="font-size:15px;font-weight:800;color:#1e40af;margin-bottom:6px;">USA-Related South Africa CompCom Case</div>
  <div style="font-size:14px;color:#1e3a8a;">This merger review appears to involve USA-related parties or markets.</div>
  <div style="margin-top:10px;">
    <a href="{escape_html(list_url)}" target="_blank" style="color:#2563eb;text-decoration:none;font-weight:700;font-size:14px;">View Weekly Case List →</a>
  </div>
</div>
<div style="background:#f3f4f6;border-radius:6px;padding:22px 26px;">
  <div style="font-size:18px;font-weight:800;margin-bottom:16px;">Case Details</div>
  <div style="display:grid;grid-template-columns:220px 1fr;row-gap:12px;column-gap:18px;">
    <div style="font-weight:700;">Case Number:</div><div>{escape_html(case_info.get("case_number", "N/A"))}</div>
    <div style="font-weight:700;">Acquiring Firm:</div><div>{escape_html(case_info.get("primary_acquiring_firm", "N/A"))}</div>
    <div style="font-weight:700;">Target Firm:</div><div>{escape_html(case_info.get("primary_target_firm", "N/A"))}</div>
    <div style="font-weight:700;">Initial Date Filed:</div><div>{escape_html(case_info.get("initial_date_filed", "N/A"))}</div>
    <div style="font-weight:700;">Size:</div><div>{escape_html(case_info.get("size", "N/A"))}</div>
    <div style="font-weight:700;">Due Date:</div><div>{escape_html(case_info.get("due_date", "N/A"))}</div>
    <div style="font-weight:700;">Phase:</div><div>{escape_html(case_info.get("phase", "N/A"))}</div>
    <div style="font-weight:700;">Status:</div><div>{escape_html(case_info.get("status", "N/A"))}</div>
  </div>
</div>
</div>
</body>
</html>"""


def send_email_via_webhook(
    subject: str,
    html_content: str,
    case_info: Dict[str, Any],
    deal_id: Optional[str] = None,
    test_mode: bool = False,
) -> bool:
    """
    Route email:
      test_mode=True  → send_direct_email to TEST_RECIPIENT via N8N_WEBHOOK_ONLY_ME
      test_mode=False → org-aware routing via post_email_payload
    """
    try:
        payload = {
            "subject": subject,
            "html": html_content,
            "case_number": case_info.get("case_number", "N/A"),
            "primary_acquiring_firm": case_info.get("primary_acquiring_firm", "N/A"),
            "primary_target_firm": case_info.get("primary_target_firm", "N/A"),
            "initial_date_filed": case_info.get("initial_date_filed", "N/A"),
            "size": case_info.get("size", "N/A"),
            "due_date": case_info.get("due_date", "N/A"),
            "extension_date": case_info.get("extension_date", ""),
            "phase": case_info.get("phase", "N/A"),
            "status": case_info.get("status", "N/A"),
            "list_week_date": case_info.get("list_week_date", ""),
            "source_xlsx_url": case_info.get("source_xlsx_url", ""),
            "list_page_url": case_info.get("list_page_url", LIST_PAGE_URL),
            "deal_id": deal_id,
            "is_new_case": True,
            "source": "sa_compcom_cases_register",
        }
        if test_mode:
            webhook_url = os.getenv("N8N_WEBHOOK_ONLY_ME", "")
            if not webhook_url:
                logger.warning(
                    "N8N_WEBHOOK_ONLY_ME not set in .env — test email skipped"
                )
                return False
            logger.info(
                "[TEST] Sending to %s via N8N_WEBHOOK_ONLY_ME", TEST_RECIPIENT
            )
            return send_direct_email(
                [TEST_RECIPIENT], payload, webhook_url=webhook_url
            )
        return post_email_payload(payload, subject=subject)
    except Exception as e:
        logger.warning("Error sending email: %s", e)
        return False


def insert_case(collection, case_info: Dict[str, Any]) -> Optional[str]:
    try:
        result = collection.insert_one(case_info)
        return str(result.inserted_id)
    except Exception as e:
        logger.error("Error inserting case: %s", e)
        return None


def reconcile_pending_removals(
    collection,
    current_case_numbers: Set[str],
    source_xlsx_url: str,
) -> int:
    """
    Silently mark DB cases with status Pending that are absent from the
    current weekly XLSX as 'removed from pending list' and set removed_at.
    """
    if not current_case_numbers:
        logger.warning("No current case numbers; skipping pending reconciliation")
        return 0

    now_iso = utc_now_iso()
    query = {
        "status": {"$regex": r"^pending$", "$options": "i"},
        "case_number": {"$nin": list(current_case_numbers)},
    }
    update = {
        "$set": {
            "status": REMOVED_STATUS,
            "removed_at": now_iso,
            "updated_at": now_iso,
            "removed_from_xlsx_url": source_xlsx_url,
        }
    }
    try:
        result = collection.update_many(query, update)
        removed = int(result.modified_count or 0)
        logger.info(
            "Pending reconciliation: marked %s case(s) as '%s'",
            removed,
            REMOVED_STATUS,
        )
        return removed
    except Exception as e:
        logger.exception("Pending reconciliation failed: %s", e)
        return 0


def _load_rows_for_links(
    links: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], Optional[str], Set[str]]:
    """
    Download + parse XLSX links. Returns (all_rows, latest_xlsx_url, case_numbers_in_latest).
    Dedupes by case_number keeping the first (newest week) occurrence.
    """
    all_rows: List[Dict[str, Any]] = []
    seen_cases: Set[str] = set()
    latest_url: Optional[str] = None
    latest_case_numbers: Set[str] = set()

    for idx, link in enumerate(links):
        path = download_xlsx(link["url"])
        if not path:
            continue
        try:
            rows = parse_xlsx_rows(
                path,
                source_xlsx_url=link["url"],
                list_week_date=link.get("week_date") or "",
            )
        finally:
            try:
                os.remove(path)
            except OSError:
                pass

        if idx == 0:
            latest_url = link["url"]
            latest_case_numbers = {
                (r.get("case_number") or "").strip()
                for r in rows
                if (r.get("case_number") or "").strip()
            }

        for row in rows:
            cn = (row.get("case_number") or "").strip()
            if not cn or cn in seen_cases:
                continue
            seen_cases.add(cn)
            all_rows.append(row)

    return all_rows, latest_url, latest_case_numbers


def run_sa_compcom_cases_register(
    bootstrap: bool = False,
    headless: bool = True,
    test_mode: bool = False,
) -> None:
    global LOG_FILE
    LOG_FILE = refresh_log_file(logger, LOG_FILE, _get_log_file)
    run_start = datetime.now()
    error_items: List[Dict[str, Any]] = []
    new_cases: List[Dict[str, Any]] = []
    llm_match_count = 0
    regex_match_count = 0
    inserted_count = 0
    skipped_existing = 0
    parsed_count = 0
    removed_pending_count = 0

    if bootstrap:
        mode_label = "Bootstrap (DB only)"
    elif test_mode:
        mode_label = f"TEST-EMAIL → {TEST_RECIPIENT}"
    else:
        mode_label = "New case monitor"
    logger.info("=" * 60)
    logger.info("Starting South Africa CompCom Cases Register — %s", mode_label)
    logger.info("Log file: %s", LOG_FILE)
    if test_mode and not bootstrap:
        logger.info(
            "TEST-EMAIL: emails → %s via N8N_WEBHOOK_ONLY_ME", TEST_RECIPIENT
        )
    logger.info("=" * 60)

    try:
        success, message = init_mongodb_connection(ENV_PATH)
        if not success:
            collect_error(
                error_items,
                f"MongoDB connection failed: {message}",
                step="mongodb_connect",
            )
            return
        logger.info("MongoDB: %s", message)

        if not is_connected():
            collect_error(
                error_items,
                "MongoDB not connected. Exiting.",
                step="mongodb_connect",
            )
            return

        collection = get_sa_compcom_cases_collection()
        if collection is None:
            collect_error(
                error_items,
                f"Could not access '{COLLECTION_NAME}' collection. Exiting.",
                step="get_collection",
            )
            return

        ensure_indexes(collection)

        html = fetch_list_page_html(LIST_PAGE_URL, headless=headless)
        if not html:
            collect_error(
                error_items,
                "Failed to fetch CompCom list page HTML",
                step="fetch_list_page_html",
                context={"url": LIST_PAGE_URL},
            )
            return

        links = parse_weekly_xlsx_links(html)
        if not links:
            collect_error(
                error_items,
                "No weekly XLSX links found on list page",
                step="parse_weekly_xlsx_links",
                context={"url": LIST_PAGE_URL},
            )
            return

        # Normal: latest week only. Bootstrap: all weeks (newest first).
        links_to_process = links if bootstrap else links[:1]
        latest_link_url = (links_to_process[0].get("url") or "").strip()
        logger.info(
            "Processing %s weekly file(s); newest=%s",
            len(links_to_process),
            links_to_process[0].get("label"),
        )

        # Normal monitor: if this week's XLSX was already ingested, skip entirely.
        if not bootstrap and xlsx_url_already_processed(
            collection, latest_link_url
        ):
            logger.info(
                "Latest XLSX already in DB (source_xlsx_url=%s); "
                "skipping entire run",
                latest_link_url,
            )
            return

        all_rows, latest_url, latest_case_numbers = _load_rows_for_links(
            links_to_process
        )
        parsed_count = len(all_rows)
        if not all_rows:
            logger.warning("No case rows parsed from XLSX. Exiting.")
            return

        open_deals = None if bootstrap else fetch_open_deals()

        for idx, row in enumerate(all_rows, 1):
            try:
                case_number = (row.get("case_number") or "").strip()
                acquiring = (row.get("primary_acquiring_firm") or "").strip()
                target = (row.get("primary_target_firm") or "").strip()
                logger.info(
                    "[%s/%s] %s | %s → %s",
                    idx,
                    len(all_rows),
                    case_number,
                    acquiring[:40],
                    target[:40],
                )

                if case_exists(collection, case_number):
                    skipped_existing += 1
                    logger.info("Already in %s; skipping", COLLECTION_NAME)
                    continue

                now_iso = utc_now_iso()
                case_info: Dict[str, Any] = {
                    **row,
                    "created_at": now_iso,
                    "updated_at": now_iso,
                }

                if bootstrap:
                    inserted_id = insert_case(collection, case_info)
                    if inserted_id:
                        inserted_count += 1
                        backup_case = dict(case_info)
                        new_cases.append(backup_case)
                        logger.info("Bootstrap inserted (id=%s)", inserted_id)
                    else:
                        collect_error(
                            error_items,
                            "Failed to insert case",
                            step="insert_case",
                            context={"case_number": case_number},
                        )
                    continue

                try:
                    matched_deal_id = match_case_to_deal(
                        case_number,
                        acquiring,
                        target,
                        deals=open_deals,
                    )
                except Exception as e:
                    logger.exception("Deal matching error: %s", e)
                    collect_error(
                        error_items,
                        str(e),
                        step="match_case_to_deal",
                        context={"case_number": case_number},
                    )
                    matched_deal_id = None

                matched_by_regex = False
                if matched_deal_id:
                    llm_match_count += 1
                else:
                    matched_deal_id = regex_match_sa_compcom_deal(
                        acquiring, target, open_deals or []
                    )
                    if matched_deal_id:
                        matched_by_regex = True
                        regex_match_count += 1
                        logger.info(
                            "Regex fallback matched deal_id=%s", matched_deal_id
                        )

                if matched_deal_id:
                    case_info["deal_id"] = matched_deal_id
                    deal = get_deal_by_id(matched_deal_id)
                    if deal:
                        subject = build_subject("sa_compcom", "new", deal)
                        subject = apply_regex_match_subject(
                            subject, matched_by_regex
                        )
                        html_email = generate_matched_case_email_html(
                            case_info, deal
                        )
                        if not send_email_via_webhook(
                            subject,
                            html_email,
                            case_info,
                            deal_id=matched_deal_id,
                            test_mode=test_mode,
                        ):
                            collect_error(
                                error_items,
                                "Failed to send matched-case email",
                                step="send_email",
                                context={
                                    "case_number": case_number,
                                    "deal_id": matched_deal_id,
                                },
                            )
                    else:
                        logger.warning(
                            "Matched deal_id=%s but deal document not found",
                            matched_deal_id,
                        )
                        collect_error(
                            error_items,
                            "Matched deal_id but deal document not found",
                            step="fetch_matched_deal",
                            context={
                                "case_number": case_number,
                                "deal_id": matched_deal_id,
                                "matched_by_regex": matched_by_regex,
                            },
                        )
                else:
                    try:
                        details_for_llm = (
                            f"Case Number: {case_number}\n"
                            f"Acquiring Firm: {acquiring}\n"
                            f"Target Firm: {target}\n"
                            f"Size: {row.get('size', '')}\n"
                            f"Phase: {row.get('phase', '')}\n"
                            f"Status: {row.get('status', '')}\n"
                            f"Initial date filed: {row.get('initial_date_filed', '')}"
                        )
                        is_usa = bool(
                            verify_usa_relation(
                                company_details=details_for_llm,
                                case_type="South Africa CompCom",
                            )
                        )
                    except Exception as e:
                        logger.exception("USA verification error: %s", e)
                        collect_error(
                            error_items,
                            str(e),
                            step="verify_usa_relation",
                            context={"case_number": case_number},
                        )
                        is_usa = False

                    if is_usa:
                        subject = build_subject("sa_compcom", "new")
                        html_email = generate_usa_related_email_html(case_info)
                        if not send_email_via_webhook(
                            subject,
                            html_email,
                            case_info,
                            test_mode=test_mode,
                        ):
                            collect_error(
                                error_items,
                                "Failed to send USA-related email",
                                step="send_email",
                                context={"case_number": case_number},
                            )
                    else:
                        logger.info(
                            "Not matched and not USA-related; silent insert"
                        )

                inserted_id = insert_case(collection, case_info)
                if inserted_id:
                    inserted_count += 1
                    backup_case = dict(case_info)
                    backup_case.pop("_id", None)
                    new_cases.append(backup_case)
                    logger.info(
                        "Inserted into %s (id=%s)", COLLECTION_NAME, inserted_id
                    )
                else:
                    collect_error(
                        error_items,
                        "Failed to insert case",
                        step="insert_case",
                        context={"case_number": case_number},
                    )
            except Exception as e:
                logger.exception("Error processing row #%s: %s", idx, e)
                collect_error(
                    error_items,
                    str(e),
                    step="process_row",
                    context={"case_number": (row.get("case_number") or "")},
                )

        # Silent pending removal against the latest weekly list only
        if latest_url and latest_case_numbers:
            removed_pending_count = reconcile_pending_removals(
                collection,
                latest_case_numbers,
                latest_url,
            )

        if new_cases:
            try:
                with open(BACKUP_JSON, "w", encoding="utf-8") as f:
                    json.dump(new_cases, f, indent=2, ensure_ascii=False)
                logger.info(
                    "Saved %s new cases to backup JSON: %s",
                    len(new_cases),
                    BACKUP_JSON,
                )
            except Exception as e:
                logger.warning("Error writing backup JSON: %s", e)

    except Exception as e:
        logger.exception("Unhandled error in run_sa_compcom_cases_register: %s", e)
        collect_error(
            error_items,
            f"Unhandled error: {e}",
            step="run_main",
        )
    finally:
        send_error_summary(error_items, SCRIPT_NAME)
        elapsed = round((datetime.now() - run_start).total_seconds(), 1)
        logger.info("=" * 60)
        logger.info("SUMMARY")
        logger.info("  Mode                         : %s", mode_label)
        logger.info("  Rows parsed                  : %s", parsed_count)
        logger.info("  Skipped (already in DB)      : %s", skipped_existing)
        logger.info("  Inserted                     : %s", inserted_count)
        if not bootstrap:
            logger.info("  LLM deal matches             : %s", llm_match_count)
            logger.info("  Regex fallback matches       : %s", regex_match_count)
        logger.info("  Pending → removed (silent)   : %s", removed_pending_count)
        logger.info("  Errors encountered           : %s", len(error_items))
        logger.info("  Total time                   : %ss", elapsed)
        logger.info("=" * 60)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="South Africa CompCom weekly case list scraper"
    )
    parser.add_argument(
        "--bootstrap",
        action="store_true",
        help="Initial import: insert all records without deal match or USA checks",
    )
    parser.add_argument(
        "--test-email",
        action="store_true",
        help=f"Send emails to {TEST_RECIPIENT} via N8N_WEBHOOK_ONLY_ME",
    )
    parser.add_argument(
        "--headed",
        action="store_true",
        help="Run Playwright with a visible browser",
    )
    args = parser.parse_args()
    run_sa_compcom_cases_register(
        bootstrap=args.bootstrap,
        headless=not args.headed,
        test_mode=args.test_email,
    )
